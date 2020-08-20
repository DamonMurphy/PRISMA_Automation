# -*- coding: utf-8 -*-
"""
Created on Wed Aug 19 08:47:49 2020

@author: murphd26
"""

from dateutil.relativedelta import relativedelta
import pandas as pd
import numpy as np
from datetime import date, datetime
import calendar
import os, sys
from os import listdir
from os.path import isfile
import shutil
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.styles import Alignment


directory_PRISMA = r'\\Ohcolnas0250\PSHP\PRISMA_Files'

os.chdir(directory_PRISMA)


# LOAD INPUT CHANNELS AND DISB CODES FILE
os.chdir('InputChannel')


workbookTitle = 'InputChannelMaster.xlsx'
inChan = pd.read_excel(workbookTitle,sheet_name = 'InputChannels',usecols='A:B').astype('str')
inChan['InputChannel'] = inChan['InputChannel'].str.upper()


#CONVERT INPUT CHANNEL DATAFRAMES TO DICTIONARY
dict_InChan = inChan.to_dict('split')
tuple_InChan = dict_InChan['data']
inChan_dict = dict(tuple_InChan)


# READ CURRENT MONTH PRISMA DATA
prisma_period = (date.today() + relativedelta(months=-1)).replace(day=1)
prisma_period = pd.Timestamp(prisma_period)

# DETERMINE FOLDER LOCATION
prisma_month = prisma_period.month
prisma_year = prisma_period.year
prisma_month_abbr = calendar.month_abbr[prisma_month]
prisma_month_name = calendar.month_name[prisma_month]


os.chdir(directory_PRISMA)
os.chdir(str(prisma_year))


year_path = os.getcwd()
dir_list = [name for name in os.listdir(year_path) if os.path.isdir(os.path.join(year_path, name))]

dir_name = [dir1 for dir1 in dir_list if (prisma_month_abbr in dir1 or prisma_month_name in dir1)]

if len(dir_name) != 1:
    print('\n\nThere are ' + str(len(dir_name)) + ' names in the folder list for the reporting month.')
    print('Check',year_path,'for correct folder then continue when fixed.\n\n')
    print('Quiting...\n\n')
    os.exit(0)

dir_name = dir_name[0]

os.chdir(dir_name)

month_path = os.getcwd()

csv_file = 'CSA_CustomPrintedJobs2-CSV-NationWide.csv'

df = pd.read_csv(csv_file)

# DROP 'device' ENTRIES WITH THE STRING 'DUM'
df = df[~df['device'].str.upper().str.contains('DUM')].reset_index(drop=True)

# RENAME COLUMNS
df.rename(columns={'jobid':'JobID','jobname':'JobName','device':'PrinterID',\
            'jescode':'JES_Code','inputchannel':'InputChannel',\
            'status':'Status','print_finished_date':'PrintDate',\
            'pages':'Images'},inplace=True)


# KEEP PERTINENT COLUMNS ONLY
df = df[['JobID','JobName','PrinterID','JES_Code','InputChannel','Status',\
       'PrintDate','Images']]


# CHANGE InputChannel TO ALL CAPS FOR MAPPING TO DisbCode
df['InputChannel'] = df['InputChannel'].str.upper()


# CONVERT JobID to STRING
df['JobID'] = df['JobID'].astype('int').astype('str')


# CONVERT JES_Code NaN to 0, then to STRING, THEN REPLACE '0' TO '0000'
# STOPS IF THERE ARE ANY INITIAL JES_Code VALUES
df['JES_Code'] = df['JES_Code'].fillna(0)
df['JES_Code'] = df['JES_Code'].astype('int').astype('str')
df['JES_Code'] = df['JES_Code'].replace(to_replace='0',value='0000')
if (df['JES_Code'].str.len() > 4).any():
    print('\nThere are JES_Code entries longer than 4 characters.')
    print('Investigate and/or fix.')
    print('Start over or change python code.\n')
    print('JES_Code entries longer than 4 characters:')
    print(df[df['JES_Code'].str.len() > 4].drop_duplicates(keep='first'))
    print('\n\nQuiting...\n\n')
    sys.exit(0)


# THE REPLACE Jes_Code WITH LEFT(3)
df['JES_Code'] = df['JES_Code'].str[:-1]


# CONVERT PrintDate TO DATETIME USING EUROPEAN FORMAT (dd.mm.yyyy)
df['PrintDate'] = pd.to_datetime(df['PrintDate'],format='%d.%m.%Y')


# CATEGORIZE PRINT BILLING CATEGORY BASED ON 8325 AND 8326
# IF 'PRT9' IS PRESENT, THEN 8325, ELSE 8326
df['Category'] = ['VPOM 8325' if 'PRT9' in printer1 else 'VPOM 8326' for printer1 in df['PrinterID']]


# CATEGORIZE PRINT BILLING CATEGORY TO MAINFRAME WHERE JES_Code NOT '000'
df.loc[df.JES_Code != '000', 'Category'] = 'Mainframe'



# APPLY INPUT CHANNEL DICTIONARY TO DETERMINE DISB CODE
df['DisbCode'] = df['InputChannel'].map(inChan_dict)


# REPLACE NaN WITH 'TBD'
df['DisbCode'] = df['DisbCode'].fillna('TBD')


# CREATE 'TBD' INPUT CHANNEL DATAFRAME
df2 = df[['InputChannel','DisbCode']]
df2 = df2[df2['DisbCode']=='TBD'].drop_duplicates(keep='first').reset_index(drop=True)


print('\n\n')
print('\'TBD\' DisbCodes:')
print(df2)


df_Mainframe = df[['Category','JES_Code','Images']]
df_Mainframe = df_Mainframe[df_Mainframe['Category']=='Mainframe'].sort_values(by=['Category','JES_Code']).reset_index(drop=True)

df_Mainframe = df_Mainframe.groupby(['Category','JES_Code']).sum().reset_index()


df_VPOM = df[['Category','InputChannel','DisbCode','Images',]]
df_VPOM = df_VPOM[df_VPOM['Category'].str.contains('VPOM')].reset_index(drop=True)


# CREATE Images8325 and Images8326 COLUMNS
df_VPOM.loc[df_VPOM.Category == 'VPOM 8325', 'Images8325'] = df.Images
df_VPOM.loc[df_VPOM.Category == 'VPOM 8326', 'Images8326'] = df.Images

# REPLACE NaN VALUES WITH '0'
df_VPOM['Images8325'] = df_VPOM['Images8325'].fillna(0)
df_VPOM['Images8326'] = df_VPOM['Images8326'].fillna(0)

# RE-TYPE AS INT
df_VPOM['Images8325'] = df_VPOM['Images8325'].astype('int')
df_VPOM['Images8326'] = df_VPOM['Images8326'].astype('int')


# RE-CREATE df_VPOM USING ONLY InpuChannel, DisbCode, Images8325 and Images8326 AS COLUMNS IN DATAFRAME
df_VPOM = df_VPOM[['InputChannel','DisbCode','Images8325','Images8326']]


df_VPOM = df_VPOM.groupby(['InputChannel','DisbCode']).sum().reset_index()
df_VPOM = df_VPOM.sort_values(by=['InputChannel','DisbCode']).reset_index(drop=True)


# WRITE NEW MAINFRAME EXCEL FILE
# 'Mainframe' SHEET IS FULL DATA
# 'Summary' IS Images BY JES_Code

df_Mainframe2 = df[df['Category']=='Mainframe'].sort_values(by=['JES_Code','PrintDate']).reset_index(drop=True)
del df_Mainframe2['DisbCode']

file1 = 'Mainframe_' + prisma_month_abbr + str(prisma_year) + '.xlsx'
ws1 = 'Mainframe_All'
ws2 = 'Mainframe_Summary'

print()
print()
print('writing data to Mainframe xlsx file...')
print()
print()

WB = Workbook()
WB.title = file1

WS = WB.active
WS.title = ws1


print('creating Mainframe sheet...')

colNames = list(df_Mainframe2.columns)

i = 65
for colName in colNames:
    WS[chr(i)+'1'] = colName
    WS[chr(i)+'1'].font = Font(bold = True)
    if ('Date' in colName or 'Images' in colName):
        WS[chr(i)+'1'].alignment = Alignment(horizontal='right')
    i+=1

print('\nHeaders finished\n\n')

NumberOfRecords = len(df_Mainframe2)

print('\nNumberOfRecords:',NumberOfRecords)
print('\n')


for i in range(len(df_Mainframe2.columns)):
    if i % 5 == 0:
            print('column',i,'out of',len(df_Mainframe2.columns),'columns')
    for j in range(NumberOfRecords):
        WS[chr(i+65)+str(j+2)] = df_Mainframe2.iloc[j,i]
        if 'DATE' in df_Mainframe2.columns[i].upper() :
            WS[chr(i+65)+str(j+2)].number_format = 'M/D/YYYY'
        if 'IMAGES' in df_Mainframe2.columns[i].upper() :
            WS[chr(i+65)+str(j+2)].number_format = '#,###,##0'

WS.column_dimensions['A'].width = 12.00
WS.column_dimensions['B'].width = 15.00
WS.column_dimensions['C'].width = 20.00
WS.column_dimensions['D'].width = 12.00
WS.column_dimensions['E'].width = 15.00
WS.column_dimensions['F'].width = 12.00
WS.column_dimensions['G'].width = 12.00
WS.column_dimensions['H'].width = 12.00
WS.column_dimensions['I'].width = 12.00



print('Mainframe sheet complete...\n\n')
print('creating Summary sheet...')


WS = WB.create_sheet(ws2)

colNames = list(df_Mainframe.columns)

i = 65
for colName in colNames:
    WS[chr(i)+'1'] = colName
    WS[chr(i)+'1'].font = Font(bold = True)
    if ('Date' in colName or 'Images' in colName):
        WS[chr(i)+'1'].alignment = Alignment(horizontal='right')
    i+=1

print('\nHeaders finished\n\n')

NumberOfRecords = len(df_Mainframe)

print('\nNumberOfRecords:',NumberOfRecords)
print('\n')


for i in range(len(df_Mainframe.columns)):
    if i % 5 == 0:
            print('column',i,'out of',len(df_Mainframe.columns),'columns')
    for j in range(NumberOfRecords):
        WS[chr(i+65)+str(j+2)] = df_Mainframe.iloc[j,i]
        if 'DATE' in df_Mainframe.columns[i].upper() :
            WS[chr(i+65)+str(j+2)].number_format = 'M/D/YYYY'
        if 'IMAGES' in df_Mainframe.columns[i].upper() :
            WS[chr(i+65)+str(j+2)].number_format = '#,###,##0'


WS.column_dimensions['A'].width = 12.00
WS.column_dimensions['B'].width = 12.00
WS.column_dimensions['C'].width = 12.00


print('Summary sheet complete...\n\n')


print()
print()
print('closing Mainframe xlsx...')
print()
print()



print('saving file...')
WB.save(file1)
print('file saved')
print('closing file...')
WB.close()
print('file closed')
del WB


# WRITE NEW VPOM EXCEL FILE
# 'VPOM' SHEET IS FULL DATA
# 'Summary' IS Images BY DisbCode

df_VPOM_All = df[df['Category'].str.contains('VPOM')].sort_values(by=['InputChannel','DisbCode','Category','PrintDate']).reset_index(drop=True)
del df_VPOM_All['JES_Code']

df1 = df_VPOM_All.copy()


file1 = 'VPOM_' + prisma_month_abbr + str(prisma_year) + '.xlsx'
ws1 = 'VPOM_All'
ws2 = 'VPOM_Summary'

print()
print()
print('writing data to VPOM xlsx file...')
print()
print()

WB = Workbook()
WB.title = file1

WS = WB.active
WS.title = ws1


print('creating VPOM_All sheet...')

colNames = list(df1.columns)

i = 65
for colName in colNames:
    WS[chr(i)+'1'] = colName
    WS[chr(i)+'1'].font = Font(bold = True)
    if ('Date' in colName or 'Images' in colName):
        WS[chr(i)+'1'].alignment = Alignment(horizontal='right')
    i+=1

print('\nHeaders finished\n\n')

NumberOfRecords = len(df1)

print('\nNumberOfRecords:',NumberOfRecords)
print('\n')


for i in range(len(df1.columns)):
    if i % 5 == 0:
            print('column',i,'out of',len(df1.columns),'columns')
    for j in range(NumberOfRecords):
        WS[chr(i+65)+str(j+2)] = df1.iloc[j,i]
        if 'DATE' in df1.columns[i].upper() :
            WS[chr(i+65)+str(j+2)].number_format = 'M/D/YYYY'
        if 'IMAGES' in df1.columns[i].upper() :
            WS[chr(i+65)+str(j+2)].number_format = '#,###,##0'

WS.column_dimensions['A'].width = 12.00
WS.column_dimensions['B'].width = 30.00
WS.column_dimensions['C'].width = 20.00
WS.column_dimensions['D'].width = 20.00
WS.column_dimensions['E'].width = 15.00
WS.column_dimensions['F'].width = 12.00
WS.column_dimensions['G'].width = 12.00
WS.column_dimensions['H'].width = 15.00
WS.column_dimensions['I'].width = 12.00


print('VPOM_All sheet complete...\n\n')
print('creating VPOM_Summary sheet...')


df1 = df_VPOM.copy()

WS = WB.create_sheet(ws2)


colNames = list(df1.columns)

i = 65
for colName in colNames:
    WS[chr(i)+'1'] = colName
    WS[chr(i)+'1'].font = Font(bold = True)
    if ('Date' in colName or 'Images' in colName):
        WS[chr(i)+'1'].alignment = Alignment(horizontal='right')
    i+=1

print('\nHeaders finished\n\n')

NumberOfRecords = len(df1)

print('\nNumberOfRecords:',NumberOfRecords)
print('\n')


for i in range(len(df1.columns)):
    if i % 5 == 0:
            print('column',i,'out of',len(df1.columns),'columns')
    for j in range(NumberOfRecords):
        WS[chr(i+65)+str(j+2)] = df1.iloc[j,i]
        if 'DATE' in df1.columns[i].upper() :
            WS[chr(i+65)+str(j+2)].number_format = 'M/D/YYYY'
        if 'IMAGES' in df1.columns[i].upper() :
            WS[chr(i+65)+str(j+2)].number_format = '#,###,##0'


WS.column_dimensions['A'].width = 20.00
WS.column_dimensions['B'].width = 15.00
WS.column_dimensions['C'].width = 15.00
WS.column_dimensions['D'].width = 15.00


print('Summary sheet complete...\n\n')


print()
print()
print('closing VPOM xlsx...')
print()
print()



print('saving file...')
WB.save(file1)
print('file saved')
print('closing file...')
WB.close()
print('file closed')
del WB


sys.exit(0)

